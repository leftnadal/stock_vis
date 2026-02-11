"""
ETF CSV 다운로더

운용사 공식 CSV/XLSX 직링크에서 ETF Holdings 데이터를 다운로드하고 파싱합니다.
HTML 스크래핑 없이 직접 다운로드 방식으로 사이트 구조 변경에 안전합니다.

지원 형식:
- CSV (ARK, iShares)
- XLSX (SPDR/State Street)

404/403 발생 시 자동 URL 복구:
1. 운용사 Holdings 페이지에서 패턴 매칭
2. LLM (Gemini Flash) 분석으로 폴백
3. 새 URL 검증 및 자동 업데이트

비용: 다운로드 $0 / URL 복구 시 LLM 호출당 ~$0.001
"""
import csv
import hashlib
import io
import logging
import re
import time
from datetime import date
from decimal import Decimal
from typing import Dict, List, Optional, Tuple

import httpx
from django.utils import timezone

from serverless.models import ETFProfile, ETFHolding


# 브라우저 헤더 (Cloudflare 우회용)
BROWSER_HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.9',
    'Accept-Encoding': 'gzip, deflate, br',
    'Connection': 'keep-alive',
    'Upgrade-Insecure-Requests': '1',
    'Sec-Fetch-Dest': 'document',
    'Sec-Fetch-Mode': 'navigate',
    'Sec-Fetch-Site': 'none',
    'Sec-Fetch-User': '?1',
    'Cache-Control': 'max-age=0',
}


logger = logging.getLogger(__name__)


# ETF CSV 소스 설정
# 실제 CSV URL은 운용사 정책에 따라 변경될 수 있음
ETF_CSV_SOURCES = {
    # ========================================
    # Tier 1: 섹터 ETF (State Street SPDR)
    # ========================================
    'XLK': {
        'name': 'Technology Select Sector SPDR Fund',
        'tier': 'sector',
        'theme_id': 'technology',
        'parser': 'spdr',
        'csv_url': 'https://www.ssga.com/us/en/intermediary/etfs/library-content/products/fund-data/etfs/us/holdings-daily-us-en-xlk.xlsx',
    },
    'XLV': {
        'name': 'Health Care Select Sector SPDR Fund',
        'tier': 'sector',
        'theme_id': 'healthcare',
        'parser': 'spdr',
        'csv_url': 'https://www.ssga.com/us/en/intermediary/etfs/library-content/products/fund-data/etfs/us/holdings-daily-us-en-xlv.xlsx',
    },
    'XLF': {
        'name': 'Financial Select Sector SPDR Fund',
        'tier': 'sector',
        'theme_id': 'financials',
        'parser': 'spdr',
        'csv_url': 'https://www.ssga.com/us/en/intermediary/etfs/library-content/products/fund-data/etfs/us/holdings-daily-us-en-xlf.xlsx',
    },
    'XLE': {
        'name': 'Energy Select Sector SPDR Fund',
        'tier': 'sector',
        'theme_id': 'energy',
        'parser': 'spdr',
        'csv_url': 'https://www.ssga.com/us/en/intermediary/etfs/library-content/products/fund-data/etfs/us/holdings-daily-us-en-xle.xlsx',
    },
    'XLI': {
        'name': 'Industrial Select Sector SPDR Fund',
        'tier': 'sector',
        'theme_id': 'industrials',
        'parser': 'spdr',
        'csv_url': 'https://www.ssga.com/us/en/intermediary/etfs/library-content/products/fund-data/etfs/us/holdings-daily-us-en-xli.xlsx',
    },
    'XLC': {
        'name': 'Communication Services Select Sector SPDR Fund',
        'tier': 'sector',
        'theme_id': 'communication',
        'parser': 'spdr',
        'csv_url': 'https://www.ssga.com/us/en/intermediary/etfs/library-content/products/fund-data/etfs/us/holdings-daily-us-en-xlc.xlsx',
    },
    'XLY': {
        'name': 'Consumer Discretionary Select Sector SPDR Fund',
        'tier': 'sector',
        'theme_id': 'consumer_discretionary',
        'parser': 'spdr',
        'csv_url': 'https://www.ssga.com/us/en/intermediary/etfs/library-content/products/fund-data/etfs/us/holdings-daily-us-en-xly.xlsx',
    },
    'XLP': {
        'name': 'Consumer Staples Select Sector SPDR Fund',
        'tier': 'sector',
        'theme_id': 'consumer_staples',
        'parser': 'spdr',
        'csv_url': 'https://www.ssga.com/us/en/intermediary/etfs/library-content/products/fund-data/etfs/us/holdings-daily-us-en-xlp.xlsx',
    },
    'XLU': {
        'name': 'Utilities Select Sector SPDR Fund',
        'tier': 'sector',
        'theme_id': 'utilities',
        'parser': 'spdr',
        'csv_url': 'https://www.ssga.com/us/en/intermediary/etfs/library-content/products/fund-data/etfs/us/holdings-daily-us-en-xlu.xlsx',
    },
    'XLRE': {
        'name': 'Real Estate Select Sector SPDR Fund',
        'tier': 'sector',
        'theme_id': 'real_estate',
        'parser': 'spdr',
        'csv_url': 'https://www.ssga.com/us/en/intermediary/etfs/library-content/products/fund-data/etfs/us/holdings-daily-us-en-xlre.xlsx',
    },
    'XLB': {
        'name': 'Materials Select Sector SPDR Fund',
        'tier': 'sector',
        'theme_id': 'materials',
        'parser': 'spdr',
        'csv_url': 'https://www.ssga.com/us/en/intermediary/etfs/library-content/products/fund-data/etfs/us/holdings-daily-us-en-xlb.xlsx',
    },

    # ========================================
    # Tier 2: 테마 ETF
    # ========================================
    'SOXX': {
        'name': 'iShares Semiconductor ETF',
        'tier': 'theme',
        'theme_id': 'semiconductor',
        'parser': 'ishares',
        'csv_url': 'https://www.ishares.com/us/products/239705/ishares-phlx-semiconductor-etf/1467271812596.ajax?fileType=csv&fileName=SOXX_holdings&dataType=fund',
    },
    'ARKK': {
        'name': 'ARK Innovation ETF',
        'tier': 'theme',
        'theme_id': 'innovation',
        'parser': 'ark',
        'csv_url': 'https://ark-funds.com/wp-content/uploads/funds-etf-csv/ARK_INNOVATION_ETF_ARKK_HOLDINGS.csv',
    },
    'ARKG': {
        'name': 'ARK Genomic Revolution ETF',
        'tier': 'theme',
        'theme_id': 'genomics',
        'parser': 'ark',
        'csv_url': 'https://ark-funds.com/wp-content/uploads/funds-etf-csv/ARK_GENOMIC_REVOLUTION_ETF_ARKG_HOLDINGS.csv',
    },
    'BOTZ': {
        'name': 'Global X Robotics & Artificial Intelligence ETF',
        'tier': 'theme',
        'theme_id': 'robotics_ai',
        'parser': 'globalx',
        'csv_url': 'https://www.globalxetfs.com/funds/botz/',  # GlobalX는 직접 CSV 없음, 예시
    },
    'TAN': {
        'name': 'Invesco Solar ETF',
        'tier': 'theme',
        'theme_id': 'solar',
        'parser': 'invesco',
        'csv_url': 'https://www.invesco.com/us/financial-products/etfs/holdings/tan',
    },
    'HACK': {
        'name': 'ETFMG Prime Cyber Security ETF',
        'tier': 'theme',
        'theme_id': 'cybersecurity',
        'parser': 'generic',
        'csv_url': '',
    },
    'LIT': {
        'name': 'Global X Lithium & Battery Tech ETF',
        'tier': 'theme',
        'theme_id': 'lithium_battery',
        'parser': 'globalx',
        'csv_url': 'https://www.globalxetfs.com/funds/lit/',
    },
    'ICLN': {
        'name': 'iShares Global Clean Energy ETF',
        'tier': 'theme',
        'theme_id': 'clean_energy',
        'parser': 'ishares',
        'csv_url': 'https://www.ishares.com/us/products/239738/ishares-global-clean-energy-etf/1467271812596.ajax?fileType=csv&fileName=ICLN_holdings&dataType=fund',
    },
    'KWEB': {
        'name': 'KraneShares CSI China Internet ETF',
        'tier': 'theme',
        'theme_id': 'china_internet',
        'parser': 'kraneshares',
        'csv_url': '',
    },
    'BETZ': {
        'name': 'Roundhill Sports Betting & iGaming ETF',
        'tier': 'theme',
        'theme_id': 'igaming',
        'parser': 'generic',
        'csv_url': '',
    },
}


class ETFCSVDownloadError(Exception):
    """ETF CSV 다운로드 에러"""
    pass


class ETFCSVParseError(Exception):
    """ETF CSV 파싱 에러"""
    pass


class ETFCSVURLRecoveryError(Exception):
    """CSV URL 자동 복구 에러"""
    pass


class ETFCSVDownloader:
    """
    ETF CSV 다운로더

    운용사 공식 CSV에서 Holdings 데이터를 다운로드하고 파싱합니다.
    변경 감지 (해시 비교) 및 에러 처리를 포함합니다.

    Usage:
        downloader = ETFCSVDownloader()

        # 단일 ETF 다운로드
        holdings = downloader.download_holdings('XLK')

        # 전체 ETF 동기화
        results = downloader.sync_all_etfs()
    """

    def __init__(self, auto_resolve_url: bool = True, max_retries: int = 3):
        """
        Args:
            auto_resolve_url: 404 발생 시 자동 URL 복구 시도 여부
            max_retries: 최대 재시도 횟수
        """
        self.client = httpx.Client(
            timeout=60.0,
            follow_redirects=True,
            headers=BROWSER_HEADERS
        )
        self._openfigi_cache: Dict[str, str] = {}
        self._auto_resolve_url = auto_resolve_url
        self._max_retries = max_retries
        self._url_resolver = None  # Lazy initialization

    def __del__(self):
        if hasattr(self, 'client'):
            self.client.close()

    def _get_url_resolver(self):
        """CSVURLResolver 인스턴스 반환 (lazy initialization)"""
        if self._url_resolver is None:
            from .csv_url_resolver import CSVURLResolver
            self._url_resolver = CSVURLResolver()
        return self._url_resolver

    def _attempt_url_recovery(self, profile: ETFProfile) -> Optional[str]:
        """
        CSV URL 자동 복구 시도

        Args:
            profile: ETFProfile 인스턴스

        Returns:
            새 URL 또는 None
        """
        try:
            resolver = self._get_url_resolver()
            new_url = resolver.resolve_csv_url(
                etf_symbol=profile.symbol,
                parser_type=profile.parser_type,
                current_url=profile.csv_url
            )

            if new_url and new_url != profile.csv_url:
                # 프로필 업데이트
                old_url = profile.csv_url
                profile.csv_url = new_url
                profile.last_error = f"URL 자동 복구됨 (이전: {old_url[:80]}...)"
                profile.save(update_fields=['csv_url', 'last_error'])
                return new_url

            return None

        except Exception as e:
            logger.warning(f"{profile.symbol}: URL 복구 실패 - {e}")
            return None

    def _resolve_globalx_url(self, etf_symbol: str) -> Optional[str]:
        """
        GlobalX ETF의 동적 CSV URL 해결

        GlobalX URL 패턴: https://assets.globalxetfs.com/funds/holdings/{etf}_full-holdings_{YYYYMMDD}.csv

        Args:
            etf_symbol: ETF 심볼 (BOTZ, LIT 등)

        Returns:
            유효한 CSV URL 또는 None
        """
        from datetime import timedelta

        base_url = "https://assets.globalxetfs.com/funds/holdings"
        symbol_lower = etf_symbol.lower()

        # 최근 7일 시도
        for days_ago in range(7):
            check_date = date.today() - timedelta(days=days_ago)
            date_str = check_date.strftime('%Y%m%d')
            url = f"{base_url}/{symbol_lower}_full-holdings_{date_str}.csv"

            try:
                response = self.client.head(url, timeout=5)
                if response.status_code == 200:
                    logger.info(f"{etf_symbol}: GlobalX URL 발견 - {date_str}")
                    return url
            except Exception:
                continue

        logger.warning(f"{etf_symbol}: GlobalX URL을 찾을 수 없음")
        return None

    def initialize_etf_profiles(self) -> int:
        """
        ETFProfile 초기화 (최초 1회 실행)

        Returns:
            생성된 프로필 수
        """
        created_count = 0
        for symbol, config in ETF_CSV_SOURCES.items():
            profile, created = ETFProfile.objects.get_or_create(
                symbol=symbol,
                defaults={
                    'name': config['name'],
                    'tier': config['tier'],
                    'theme_id': config['theme_id'],
                    'csv_url': config.get('csv_url', ''),
                    'parser_type': config.get('parser', 'generic'),
                }
            )
            if created:
                created_count += 1
                logger.info(f"ETFProfile 생성: {symbol}")

        return created_count

    def download_holdings(self, etf_symbol: str) -> List[Dict]:
        """
        단일 ETF의 Holdings 다운로드

        Args:
            etf_symbol: ETF 심볼 (예: XLK, SOXX)

        Returns:
            Holdings 리스트:
            [
                {
                    'symbol': 'AAPL',
                    'weight': 22.5,
                    'shares': 1000000,
                    'market_value': 150000000,
                    'rank': 1
                },
                ...
            ]

        Raises:
            ETFCSVDownloadError: 다운로드 실패
            ETFCSVParseError: 파싱 실패
        """
        try:
            profile = ETFProfile.objects.get(symbol=etf_symbol.upper())
        except ETFProfile.DoesNotExist:
            raise ETFCSVDownloadError(f"ETF 프로필 없음: {etf_symbol}")

        if not profile.csv_url:
            # GlobalX 동적 URL 시도
            if profile.parser_type == 'globalx':
                dynamic_url = self._resolve_globalx_url(etf_symbol)
                if dynamic_url:
                    profile.csv_url = dynamic_url
                    profile.save(update_fields=['csv_url'])
                else:
                    raise ETFCSVDownloadError(f"CSV URL 미설정: {etf_symbol}")
            else:
                raise ETFCSVDownloadError(f"CSV URL 미설정: {etf_symbol}")

        # 다운로드 with 재시도
        content = None
        download_url = profile.csv_url

        # GlobalX는 날짜 기반 URL이므로 동적 해결
        if profile.parser_type == 'globalx' and '{date}' not in download_url:
            dynamic_url = self._resolve_globalx_url(etf_symbol)
            if dynamic_url:
                download_url = dynamic_url
        content_type = None
        last_error = None

        for attempt in range(self._max_retries):
            try:
                response = self.client.get(download_url)
                response.raise_for_status()
                content = response.content
                content_type = response.headers.get('content-type', '')
                break
            except httpx.HTTPStatusError as e:
                last_error = e
                # 404/403 에러 시 자동 URL 복구 시도 (첫 번째 시도에서만)
                if e.response.status_code in (404, 403) and self._auto_resolve_url and attempt == 0:
                    logger.warning(
                        f"{etf_symbol}: HTTP {e.response.status_code} 에러, URL 자동 복구 시도"
                    )
                    new_url = self._attempt_url_recovery(profile)
                    if new_url:
                        download_url = new_url
                        continue
                # 재시도 가능한 에러인지 확인
                if e.response.status_code in (429, 500, 502, 503, 504):
                    time.sleep(2 ** attempt)  # 지수 백오프
                    continue
                break
            except httpx.HTTPError as e:
                last_error = e
                if attempt < self._max_retries - 1:
                    time.sleep(2 ** attempt)
                    continue
                break

        if content is None:
            error_msg = f"다운로드 실패 (재시도 {self._max_retries}회): {str(last_error)}"
            profile.last_error = error_msg
            profile.save(update_fields=['last_error'])
            raise ETFCSVDownloadError(error_msg)

        # 변경 감지
        content_hash = hashlib.sha256(content).hexdigest()
        if content_hash == profile.last_hash:
            logger.info(f"{etf_symbol}: 변경 없음 (hash 동일)")
            return self._get_cached_holdings(profile)

        # 파싱 (XLSX vs CSV 자동 감지)
        try:
            is_xlsx = self._is_xlsx_content(content, content_type, download_url)
            if is_xlsx:
                holdings = self._parse_xlsx(content, profile.parser_type, etf_symbol)
            else:
                holdings = self._parse_csv(content, profile.parser_type, etf_symbol)
        except Exception as e:
            error_msg = f"파싱 실패: {str(e)}"
            profile.last_error = error_msg
            profile.save(update_fields=['last_error'])
            raise ETFCSVParseError(error_msg)

        if not holdings:
            error_msg = "파싱된 Holdings가 없습니다"
            profile.last_error = error_msg
            profile.save(update_fields=['last_error'])
            raise ETFCSVParseError(error_msg)

        # DB 저장
        self._save_holdings(profile, holdings, content_hash)

        return holdings

    def _is_xlsx_content(self, content: bytes, content_type: str, url: str) -> bool:
        """XLSX 파일인지 확인"""
        # Content-Type 체크
        xlsx_types = [
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'application/vnd.ms-excel',
        ]
        if any(t in content_type.lower() for t in xlsx_types):
            return True

        # URL 확장자 체크
        if url.lower().endswith('.xlsx') or url.lower().endswith('.xls'):
            return True

        # Magic bytes 체크 (PK = ZIP = XLSX)
        if content[:2] == b'PK':
            return True

        return False

    def _parse_xlsx(
        self,
        content: bytes,
        parser_type: str,
        etf_symbol: str
    ) -> List[Dict]:
        """
        XLSX 파일 파싱 (openpyxl 사용)

        Args:
            content: XLSX 바이트 데이터
            parser_type: 파서 타입
            etf_symbol: ETF 심볼

        Returns:
            파싱된 Holdings 리스트
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ETFCSVParseError("openpyxl 라이브러리가 설치되지 않았습니다: pip install openpyxl")

        holdings = []

        try:
            wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active

            # 모든 행을 리스트로 변환 (read_only 모드에서 iter_rows는 한번만 순회 가능)
            all_rows = list(ws.iter_rows(values_only=True))

            # 헤더 행 찾기
            header_row_idx = None
            header_map = {}

            for row_idx, row in enumerate(all_rows):
                if row is None:
                    continue
                row_str = ' '.join(str(cell or '') for cell in row).lower()

                # SPDR 형식: Name, Ticker, Weight 등
                if 'ticker' in row_str and ('weight' in row_str or 'percentage' in row_str):
                    header_row_idx = row_idx
                    for col_idx, cell in enumerate(row):
                        if cell:
                            header_map[str(cell).strip().lower()] = col_idx
                    break

            if header_row_idx is None:
                logger.warning(f"{etf_symbol}: XLSX 헤더 행을 찾을 수 없음")
                return []

            # 데이터 행 파싱 (대소문자 무시)
            ticker_col = None
            weight_col = None

            for key, col in header_map.items():
                if key in ('ticker', 'symbol'):
                    ticker_col = col
                if key in ('weight', 'weight (%)', 'percentage', 'weight(%)'):
                    weight_col = col
            shares_col = header_map.get('shares held', header_map.get('shares'))

            if ticker_col is None or weight_col is None:
                logger.warning(f"{etf_symbol}: 필수 컬럼(Ticker, Weight)을 찾을 수 없음. header_map={header_map}")
                return []

            for row in all_rows[header_row_idx + 1:]:
                if row is None or len(row) <= max(ticker_col, weight_col):
                    continue

                ticker = row[ticker_col]
                if not ticker or str(ticker).strip() in ('-', '', 'None'):
                    continue

                ticker = str(ticker).strip().upper()

                # 현금/선물 등 제외
                if any(x in ticker for x in ['CASH', 'USD', 'MARGIN', 'FUTURE', 'GOVT']):
                    continue

                try:
                    weight = self._parse_decimal(str(row[weight_col]) if row[weight_col] else '0')
                    shares = None
                    if shares_col and shares_col < len(row):
                        shares = self._parse_int(str(row[shares_col]) if row[shares_col] else '0')

                    holdings.append({
                        'symbol': ticker,
                        'weight': float(weight) if weight else 0,
                        'shares': shares,
                        'market_value': None,
                    })
                except (ValueError, TypeError) as e:
                    logger.warning(f"{etf_symbol}: 행 파싱 실패 - {row} - {e}")
                    continue

            wb.close()

        except Exception as e:
            raise ETFCSVParseError(f"XLSX 파싱 오류: {str(e)}")

        # 순위 부여 및 정렬
        holdings = sorted(holdings, key=lambda x: x['weight'], reverse=True)
        for i, h in enumerate(holdings, start=1):
            h['rank'] = i

        logger.info(f"{etf_symbol}: {len(holdings)}개 종목 파싱 완료 (XLSX)")
        return holdings

    def _parse_csv(
        self,
        content: bytes,
        parser_type: str,
        etf_symbol: str
    ) -> List[Dict]:
        """
        CSV 파싱 (운용사별 파서)

        Args:
            content: CSV 바이트 데이터
            parser_type: 파서 타입 (spdr, ishares, ark, etc.)
            etf_symbol: ETF 심볼 (로깅용)

        Returns:
            파싱된 Holdings 리스트
        """
        # 인코딩 감지 및 디코딩
        try:
            text = content.decode('utf-8')
        except UnicodeDecodeError:
            text = content.decode('latin-1')

        # 파서 선택
        parser_map = {
            'spdr': self._parse_spdr_csv,
            'ishares': self._parse_ishares_csv,
            'ark': self._parse_ark_csv,
            'globalx': self._parse_globalx_csv,
            'invesco': self._parse_invesco_csv,
            'generic': self._parse_generic_csv,
        }

        parser = parser_map.get(parser_type, self._parse_generic_csv)
        holdings = parser(text, etf_symbol)

        # 순위 부여
        for i, holding in enumerate(holdings, start=1):
            holding['rank'] = i

        logger.info(f"{etf_symbol}: {len(holdings)}개 종목 파싱 완료")
        return holdings

    def _parse_spdr_csv(self, text: str, etf_symbol: str) -> List[Dict]:
        """
        State Street SPDR CSV 파싱

        SPDR ETF CSV 형식:
        - 상단에 메타데이터 행 존재 (As of Date 등)
        - 실제 데이터는 몇 줄 아래에서 시작
        - 컬럼: Name, Ticker, Identifier, SEDOL, Weight, Sector, Shares Held, Local Currency
        """
        holdings = []
        lines = text.strip().split('\n')

        # 헤더 행 찾기
        header_idx = -1
        for i, line in enumerate(lines):
            if 'Ticker' in line and 'Weight' in line:
                header_idx = i
                break

        if header_idx == -1:
            logger.warning(f"{etf_symbol}: 헤더 행을 찾을 수 없음")
            return self._parse_generic_csv(text, etf_symbol)

        # CSV 파싱
        reader = csv.DictReader(lines[header_idx:])
        for row in reader:
            ticker = row.get('Ticker', '').strip()
            if not ticker or ticker == '-':
                continue

            try:
                weight = self._parse_decimal(row.get('Weight', '0'))
                shares = self._parse_int(row.get('Shares Held', '0'))

                holdings.append({
                    'symbol': ticker,
                    'weight': float(weight) if weight else 0,
                    'shares': shares,
                    'market_value': None,
                })
            except (ValueError, TypeError) as e:
                logger.warning(f"{etf_symbol}: 행 파싱 실패 - {row} - {e}")
                continue

        return sorted(holdings, key=lambda x: x['weight'], reverse=True)

    def _parse_ishares_csv(self, text: str, etf_symbol: str) -> List[Dict]:
        """
        iShares CSV 파싱

        iShares CSV 형식:
        - 상단에 펀드 정보 행 (9-10줄)
        - 컬럼: Ticker, Name, Sector, Asset Class, Market Value, Weight (%), Notional Value, Shares, CUSIP, ISIN, SEDOL, Price, Location, Exchange, Currency, FX Rate, Market Currency
        """
        holdings = []
        lines = text.strip().split('\n')

        # 헤더 행 찾기 (Ticker와 Weight가 모두 있는 행)
        header_idx = -1
        for i, line in enumerate(lines):
            if 'Ticker' in line and ('Weight' in line or 'Weight (%)' in line):
                header_idx = i
                break

        if header_idx == -1:
            return self._parse_generic_csv(text, etf_symbol)

        reader = csv.DictReader(lines[header_idx:])
        for row in reader:
            ticker = row.get('Ticker', '').strip().strip('"')
            if not ticker or ticker == '-':
                continue

            # 티커 길이 검증 (10자 초과 = 면책 조항 등 비정상 데이터)
            if len(ticker) > 10:
                continue

            # 현금/선물 등 제외
            if any(x in ticker.upper() for x in ['CASH', 'USD', 'MARGIN', 'FUTURE']):
                continue

            try:
                weight = self._parse_decimal(row.get('Weight (%)', row.get('Weight', '0')))
                shares = self._parse_int(row.get('Shares', '0'))
                market_value = self._parse_decimal(row.get('Market Value', '0'))

                holdings.append({
                    'symbol': ticker,
                    'weight': float(weight) if weight else 0,
                    'shares': shares,
                    'market_value': float(market_value) if market_value else None,
                })
            except (ValueError, TypeError):
                continue

        return sorted(holdings, key=lambda x: x['weight'], reverse=True)

    def _parse_ark_csv(self, text: str, etf_symbol: str) -> List[Dict]:
        """
        ARK Invest CSV 파싱

        ARK CSV 형식:
        - 컬럼: date, fund, company, ticker, cusip, shares, "market value ($)", "weight (%)"
        """
        holdings = []
        reader = csv.DictReader(io.StringIO(text))

        for row in reader:
            ticker = row.get('ticker', '').strip()
            if not ticker:
                continue

            try:
                weight = self._parse_decimal(row.get('weight (%)', '0'))
                shares = self._parse_int(row.get('shares', '0'))
                market_value = self._parse_decimal(row.get('market value ($)', '0'))

                holdings.append({
                    'symbol': ticker,
                    'weight': float(weight) if weight else 0,
                    'shares': shares,
                    'market_value': float(market_value) if market_value else None,
                })
            except (ValueError, TypeError):
                continue

        return sorted(holdings, key=lambda x: x['weight'], reverse=True)

    def _parse_globalx_csv(self, text: str, etf_symbol: str) -> List[Dict]:
        """
        Global X CSV 파싱 (BOTZ, LIT 등)

        GlobalX CSV 형식:
        - Line 1: ETF 이름
        - Line 2: "Fund Holdings Data as of MM/DD/YYYY"
        - Line 3: 헤더 "% of Net Assets,Ticker,Name,SEDOL,..."
        - Line 4+: 데이터
        """
        holdings = []
        lines = text.strip().split('\n')

        # 헤더 행 찾기 (% of Net Assets 포함)
        header_idx = -1
        for i, line in enumerate(lines):
            if '% of Net Assets' in line or 'Ticker' in line:
                header_idx = i
                break

        if header_idx == -1:
            return self._parse_generic_csv(text, etf_symbol)

        reader = csv.DictReader(lines[header_idx:])
        for row in reader:
            ticker_raw = row.get('Ticker')
            if ticker_raw is None:
                continue
            ticker = str(ticker_raw).strip().strip('"')
            if not ticker or ticker == '-' or ticker == '':
                continue

            # 티커 길이 검증 (홍콩/해외 주식 제외)
            if len(ticker) > 10 or ' ' in ticker:
                continue

            # 현금/선물 등 제외
            if any(x in ticker.upper() for x in ['CASH', 'USD', 'MARGIN', 'FUTURE']):
                continue

            try:
                weight = self._parse_decimal(str(row.get('% of Net Assets') or '0'))
                shares = self._parse_int(str(row.get('Shares Held') or '0'))
                market_value = self._parse_decimal(str(row.get('Market Value ($)') or '0'))

                holdings.append({
                    'symbol': ticker,
                    'weight': float(weight) if weight else 0,
                    'shares': shares,
                    'market_value': float(market_value) if market_value else None,
                })
            except (ValueError, TypeError):
                continue

        return sorted(holdings, key=lambda x: x['weight'], reverse=True)

    def _parse_invesco_csv(self, text: str, etf_symbol: str) -> List[Dict]:
        """
        Invesco CSV 파싱 (TAN 등)

        Note: Invesco도 직접 CSV 제공하지 않음
              현재는 generic fallback
        """
        return self._parse_generic_csv(text, etf_symbol)

    def _parse_generic_csv(self, text: str, etf_symbol: str) -> List[Dict]:
        """
        범용 CSV 파서

        일반적인 Holdings CSV 형식 추정:
        - 헤더에 Ticker/Symbol과 Weight/Percent 포함
        """
        holdings = []
        lines = text.strip().split('\n')

        # 헤더 찾기
        header_idx = 0
        for i, line in enumerate(lines):
            lower = line.lower()
            if ('ticker' in lower or 'symbol' in lower) and ('weight' in lower or 'percent' in lower):
                header_idx = i
                break

        try:
            reader = csv.DictReader(lines[header_idx:])
            # 컬럼명 정규화
            for row in reader:
                # 티커 찾기
                ticker = None
                for key in ['Ticker', 'ticker', 'Symbol', 'symbol', 'TICKER', 'SYMBOL']:
                    if key in row:
                        ticker = row[key].strip()
                        break

                if not ticker or ticker == '-':
                    continue

                # Weight 찾기
                weight = 0
                for key in ['Weight', 'weight', 'Weight (%)', 'weight (%)', 'Percent', 'percent', '%']:
                    if key in row:
                        weight = self._parse_decimal(row[key])
                        break

                holdings.append({
                    'symbol': ticker,
                    'weight': float(weight) if weight else 0,
                    'shares': None,
                    'market_value': None,
                })
        except Exception as e:
            logger.warning(f"{etf_symbol}: Generic 파싱 실패 - {e}")

        return sorted(holdings, key=lambda x: x['weight'], reverse=True)

    def _parse_decimal(self, value: str) -> Optional[Decimal]:
        """문자열을 Decimal로 변환"""
        if not value:
            return None
        # 쉼표, 퍼센트 기호 제거
        cleaned = re.sub(r'[,%$]', '', str(value).strip())
        if not cleaned or cleaned == '-':
            return None
        try:
            return Decimal(cleaned)
        except Exception:
            return None

    def _parse_int(self, value: str) -> Optional[int]:
        """문자열을 정수로 변환"""
        if not value:
            return None
        cleaned = re.sub(r'[,]', '', str(value).strip())
        if not cleaned or cleaned == '-':
            return None
        try:
            return int(float(cleaned))
        except Exception:
            return None

    def _save_holdings(
        self,
        profile: ETFProfile,
        holdings: List[Dict],
        content_hash: str
    ) -> int:
        """
        Holdings 데이터 DB 저장

        기존 데이터 삭제 후 새 데이터 저장 (전체 교체)
        중복 티커는 비중을 합산하여 저장

        Args:
            profile: ETFProfile 인스턴스
            holdings: 파싱된 Holdings 리스트
            content_hash: CSV 콘텐츠 해시

        Returns:
            저장된 레코드 수
        """
        today = date.today()

        # 오늘자 기존 데이터 삭제
        ETFHolding.objects.filter(etf=profile, snapshot_date=today).delete()

        # 중복 티커 합산
        deduped = {}
        for h in holdings:
            symbol = h['symbol']
            if symbol in deduped:
                # 비중 합산
                deduped[symbol]['weight'] += h['weight']
                # shares와 market_value도 합산
                if h.get('shares') and deduped[symbol].get('shares'):
                    deduped[symbol]['shares'] += h['shares']
                if h.get('market_value') and deduped[symbol].get('market_value'):
                    deduped[symbol]['market_value'] += h['market_value']
            else:
                deduped[symbol] = h.copy()

        # 비중 순으로 재정렬 및 순위 부여
        holdings_list = sorted(deduped.values(), key=lambda x: x['weight'], reverse=True)
        for i, h in enumerate(holdings_list, start=1):
            h['rank'] = i

        # 새 데이터 저장
        holding_objects = []
        for h in holdings_list:
            holding_objects.append(ETFHolding(
                etf=profile,
                stock_symbol=h['symbol'],
                weight_percent=Decimal(str(h['weight'])),
                shares=h.get('shares'),
                market_value=Decimal(str(h['market_value'])) if h.get('market_value') else None,
                rank=h['rank'],
                snapshot_date=today,
            ))

        ETFHolding.objects.bulk_create(holding_objects)

        # 프로필 업데이트
        profile.last_updated = timezone.now()
        profile.last_row_count = len(holdings)
        profile.last_hash = content_hash
        profile.last_error = ''
        profile.save(update_fields=['last_updated', 'last_row_count', 'last_hash', 'last_error'])

        logger.info(f"{profile.symbol}: {len(holdings)}개 Holdings 저장 완료")
        return len(holdings)

    def _get_cached_holdings(self, profile: ETFProfile) -> List[Dict]:
        """
        DB에서 캐시된 Holdings 조회

        Args:
            profile: ETFProfile 인스턴스

        Returns:
            Holdings 리스트
        """
        holdings = ETFHolding.objects.filter(etf=profile).order_by('rank')
        return [
            {
                'symbol': h.stock_symbol,
                'weight': float(h.weight_percent),
                'shares': h.shares,
                'market_value': float(h.market_value) if h.market_value else None,
                'rank': h.rank,
            }
            for h in holdings
        ]

    def sync_all_etfs(self, tier: Optional[str] = None) -> Dict[str, Dict]:
        """
        전체 ETF Holdings 동기화

        Args:
            tier: 특정 tier만 동기화 ('sector' 또는 'theme'), None이면 전체

        Returns:
            {
                'XLK': {'status': 'success', 'count': 75},
                'SOXX': {'status': 'failed', 'error': '...'},
                ...
            }
        """
        profiles = ETFProfile.objects.filter(is_active=True)
        if tier:
            profiles = profiles.filter(tier=tier)

        results = {}
        for profile in profiles:
            try:
                holdings = self.download_holdings(profile.symbol)
                results[profile.symbol] = {
                    'status': 'success',
                    'count': len(holdings),
                    'last_updated': profile.last_updated.isoformat() if profile.last_updated else None,
                }
            except (ETFCSVDownloadError, ETFCSVParseError) as e:
                results[profile.symbol] = {
                    'status': 'failed',
                    'error': str(e),
                }
            except Exception as e:
                results[profile.symbol] = {
                    'status': 'error',
                    'error': f"Unexpected: {str(e)}",
                }

        return results

    def map_cusip_to_ticker(self, cusip: str) -> Optional[str]:
        """
        CUSIP → Ticker 매핑 (OpenFIGI API 사용)

        Args:
            cusip: CUSIP 코드

        Returns:
            티커 심볼 (없으면 None)

        Note: OpenFIGI는 무료지만 Rate limit 있음 (25 req/min)
        """
        if cusip in self._openfigi_cache:
            return self._openfigi_cache[cusip]

        try:
            response = self.client.post(
                'https://api.openfigi.com/v3/mapping',
                json=[{'idType': 'ID_CUSIP', 'idValue': cusip}],
                headers={'Content-Type': 'application/json'}
            )
            response.raise_for_status()
            data = response.json()

            if data and len(data) > 0 and 'data' in data[0]:
                ticker = data[0]['data'][0].get('ticker')
                self._openfigi_cache[cusip] = ticker
                return ticker

        except Exception as e:
            logger.warning(f"OpenFIGI 매핑 실패 (CUSIP: {cusip}): {e}")

        return None

    def detect_changes(self, etf_symbol: str) -> Tuple[bool, int, int]:
        """
        ETF Holdings 변경 감지 (실제 다운로드 없이)

        Args:
            etf_symbol: ETF 심볼

        Returns:
            (변경 여부, 이전 row count, 현재 row count)

        Note: HEAD 요청 + Content-Length 비교로 변경 추정
        """
        try:
            profile = ETFProfile.objects.get(symbol=etf_symbol.upper())
        except ETFProfile.DoesNotExist:
            return False, 0, 0

        if not profile.csv_url:
            return False, profile.last_row_count, 0

        try:
            response = self.client.head(profile.csv_url)
            content_length = int(response.headers.get('Content-Length', 0))
            # 대략적인 row 추정 (평균 100바이트/row)
            estimated_rows = content_length // 100

            # 이전 row count와 10% 이상 차이나면 변경으로 판단
            prev_count = profile.last_row_count or 0
            if prev_count == 0:
                return True, prev_count, estimated_rows

            diff_ratio = abs(estimated_rows - prev_count) / prev_count
            return diff_ratio > 0.1, prev_count, estimated_rows

        except Exception:
            return False, profile.last_row_count or 0, 0


# 싱글톤 인스턴스
_downloader_instance: Optional[ETFCSVDownloader] = None


def get_etf_csv_downloader() -> ETFCSVDownloader:
    """ETFCSVDownloader 싱글톤 인스턴스 반환"""
    global _downloader_instance
    if _downloader_instance is None:
        _downloader_instance = ETFCSVDownloader()
    return _downloader_instance
